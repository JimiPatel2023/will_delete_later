import re
import os
import logging
import pandas as pd
from openpyxl import load_workbook

file_path = "Scraped_Prices.xlsx"

class ExcelManipulation:
    def tensorflow(self):
        os.environ['OMP_NUM_THREADS'] = '1'

    def count_rows_in_excel(self):
        df = pd.read_excel(file_path)
        return len(df)

    def extract_prices(self, price_range):
        prices = re.findall(r'£([0-9]+(?:\.[0-9]{1,2})?)', price_range)
        prices = [float(price) for price in prices]
        if len(prices) == 1:
            prices.append(0.00)
        if len(prices) == 0:
            prices.append(0.00)
            prices.append(0.00)

        return prices[0], prices[1]


    def append_to_excel(self, item, url, price1, price2):
        workbook = load_workbook(file_path)
        sheet = workbook.active
        headers = ["URL", "Item", "Price1", "Price2"]
        next_row = sheet.max_row + 1
        sheet[f"A{next_row}"] = item
        sheet[f"B{next_row}"] = url
        sheet[f"C{next_row}"] = price1
        sheet[f"D{next_row}"] = price2
        workbook.save(file_path)

    def remove_duplicates_in_url(self):
        df = pd.read_excel(file_path)
        df_cleaned = df.drop_duplicates(subset='URL', keep='first')
        df_cleaned.to_excel(file_path, index=False)

    def read_excel_rows_as_list(self):
        df = pd.read_excel(file_path)
        rows_as_list = df[['Item', 'URL', 'Price1', 'Price2']].values.tolist()
        return rows_as_list

    def calculate_percentage_drop(self, oldPrice1, oldPrice2, newPrice1, newPrice2):
        drop_price1 = 0
        drop_price2 = 0
        oldPrice1, oldPrice2, newPrice1, newPrice2 = float(oldPrice1), float(oldPrice2), float(newPrice1), float(newPrice2)
        def percentage_drop(old_price, new_price):
            if old_price != 0 and new_price < old_price:
                drop = ((old_price - new_price) / old_price) * 100
                return round(drop, 2)
            return 0

        if oldPrice1 != 0 and newPrice1:
            drop_price1 = percentage_drop(oldPrice1, newPrice1)

        if oldPrice2 != 0 and newPrice2:
            drop_price2 = percentage_drop(oldPrice2, newPrice2)

        return drop_price1, drop_price2

    def update_prices(self, item, newPrice1, newPrice2):
        df = pd.read_excel(file_path)

        if item in df['Item'].values:
            if int(newPrice1)!=0:
                df.loc[df['Item'] == item, 'Price1'] = newPrice1
            if int(newPrice2) != 0:
                df.loc[df['Item'] == item, 'Price2'] = newPrice2

            df.to_excel(file_path, index=False)
            print(f"Prices updated for item: {item}")
        else:
            print(f"Item '{item}' not found in the Excel sheet.")

    def extract_float(self, price_str):
        try:
            price_float = float(price_str.replace('£', '').strip())
        except:
            price_float = 0.0
        return price_float
